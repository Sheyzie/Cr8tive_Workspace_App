import csv
import os
import pdfplumber
from openpyxl import load_workbook
from exceptions.exception import ValidationError


class ImportManager:
    def __init__(self, file_path=None, file_type=None, has_header=False):
        self.file_path = file_path
        self.file_type = file_type.lower() if file_type else None
        self.has_header = has_header
        self._validate()

    def _validate(self):
        ACCEPTED_TYPES = {'.csv', '.xls', '.xlsx', '.pdf'}

        if not self.file_path:
            raise ValidationError('File path is required.')

        if not self.file_type:
            raise ValidationError('File type is required.')


        if not os.path.exists(self.file_path):
            raise ValidationError('File path does not exist')
        
        if self.file_type not in ACCEPTED_TYPES:
            raise ValidationError(f'File type: {self.file_type} not valid. Most include {', '.join(ACCEPTED_TYPES)}')
        
    def import_from_csv(self):
        if self.file_type != '.csv':
            raise ValidationError('Incorrect file type. Expected a CSV file.')
    
        if not self.has_header:
            with open(self.file_path, mode='r', newline='') as file:
                reader = csv.reader(file)
                for row in reader:
                    yield row
        else:
            with open(self.file_path, mode='r', newline='') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    yield row

    def import_from_excel(self):
        if self.file_type not in {'.xls', '.xlsx'}:
            raise ValidationError('Incorrect file type. Expected a XLS or XLSX file.')
        
        wb = load_workbook(self.file_path)
        sheet = wb.active
        start_row = sheet.min_row + 1 if self.has_header else sheet.min_row
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
            data = [
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value if row[4].value else ""
            ]
        
            yield data

    def import_from_pdf(self):
        if self.file_type != '.pdf':
            raise ValidationError('Incorrect file type. Expected a PDF file.')
        
        with pdfplumber.open(self.file_path) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
            
                if table:
                    for row in table:
                        yield row